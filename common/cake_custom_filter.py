"""
CAKE by VPBank — Social Listening Spam Filter
==============================================
Phân loại bài đăng liên quan đến CAKE (VPBank fintech) vs bánh thật / nội dung không liên quan.

Logic 3 bước:
  1. FINTECH keywords mạnh  → is_spam = NO  (CAKE_FINTECH)
  2. BAKERY keywords mạnh   → is_spam = YES (BAKERY)
  3. Nội dung không liên quan → is_spam = YES (UNRELATED)

Usage:
    python cake_spam_filter.py --input data.xlsx --output result.xlsx
    python cake_spam_filter.py --input data.xlsx  # output = data_filtered.xlsx
"""

import argparse
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


# =============================================================================
# KEYWORD LISTS — chỉnh sửa tại đây để cập nhật rules
# =============================================================================

# Nhóm 1: CAKE VPBank fintech → KHÔNG phải spam
FINTECH_STRONG = [
    "vpbank", "cakebyvpbank", "#cakebyvpbank", "#proudofcake",
    "huy động", "số dư", "bi team", "fintech", "kỳ lân", "unicorn",
    "triệu user", "triệu khách hàng", "nghìn tỷ", "15.000 tỷ",
    "10.000 tỷ", "1.000 tỷ", "app mới launch", "cake app",
    "team cake", "cả team cake", "sản phẩm ở cake", "happy birthday cake !!!",
]

# Nhóm 2: Tiệm bánh / ngành bánh → SPAM (BAKERY)
BAKERY_STRONG = [
    "tiệm bánh", "đặt bánh", "giao bánh", "ship bánh", "bánh kem",
    "bánh sinh nhật", "thổi nến", "lấy gấp", "cốt bánh",
    "kem sữa tươi", "bông lan", "mousse", "tiramisu", "panna cotta",
    "bánh rút tiền", "gato", "gatô", "entremet", "cheesecake",
    "bánh bento", "bánh mousse", "bánh vẽ", "bánh in ảnh",
    "bánh cưới", "bánh thôi nôi", "bánh mừng thọ",
    "nhận đặt bánh", "order bánh", "ship tận nơi",
    "thanh toán khi nhận", "lấy ngay",
    "#banhsinhnhat", "#banhkem", "#tiembanhshincake",
    "#makicake", "baking is love", "birthday cake",
    "wedding cake", "butter cake", "chocolate cake",
    "fruit cake", "marble cake", "cup cake",
]

# Nhóm 3: Nội dung hoàn toàn không liên quan → SPAM (UNRELATED)
OTHER_SPAM = [
    "hải sản", "buffet", "huda beauty", "phấn phủ", "mỹ phẩm",
    "yến sào", "khánh hòa", "gozyuger", "tokusatsu", "super sentai",
    "anime", "kamen rider", "highlands coffee", "voucher",
    "party shop", "balloon", "decoration", "tuyển dụng", "kế toán",
    "vốn kinh doanh", "rice cake shop", "drug rice cake",
    "flipbook", "công thức bánh", "cheft",
    "srilanka", "rathnapura", "chilaw",
    "bake loose", "setting powder", "loose baking",
]


# =============================================================================
# CLASSIFICATION LOGIC
# =============================================================================

def classify_row(row: pd.Series) -> tuple[str, str]:
    """
    Phân loại một dòng dữ liệu dựa trên Title + Content + Description.

    Returns:
        (is_spam, spam_reason)
        is_spam      : "YES" hoặc "NO"
        spam_reason  : "CAKE_FINTECH" | "BAKERY" | "UNRELATED" | "UNKNOWN"
    """
    text = " ".join([
        str(row.get("Title", "") or ""),
        str(row.get("Content", "") or ""),
        str(row.get("Description", "") or ""),
    ]).lower()

    # Bước 1: FINTECH → không spam
    for kw in FINTECH_STRONG:
        if kw.lower() in text:
            return "NO", "CAKE_FINTECH"

    # Bước 2: BAKERY → spam
    for kw in BAKERY_STRONG:
        if kw.lower() in text:
            return "YES", "BAKERY"

    # Bước 3: Nội dung khác → spam
    for kw in OTHER_SPAM:
        if kw.lower() in text:
            return "YES", "UNRELATED"

    return "NO", "UNKNOWN"


def run_filter(df: pd.DataFrame) -> pd.DataFrame:
    """Áp dụng bộ filter lên toàn bộ DataFrame, thêm cột is_spam & spam_reason."""
    results = df.apply(classify_row, axis=1, result_type="expand")
    df = df.copy()
    df["is_spam"] = results[0]
    df["spam_reason"] = results[1]
    return df


# =============================================================================
# EXCEL OUTPUT WITH COLOR FORMATTING
# =============================================================================

def save_colored_excel(df: pd.DataFrame, output_path: str) -> None:
    """Lưu DataFrame ra file Excel với màu sắc theo kết quả phân loại."""
    df.to_excel(output_path, index=False, engine="openpyxl")

    wb = load_workbook(output_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    spam_col   = headers.index("is_spam") + 1
    reason_col = headers.index("spam_reason") + 1

    # Header row style
    header_fill = PatternFill("solid", fgColor="2C3E50")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Row colors
    red_fill   = PatternFill("solid", fgColor="FDDEDE")  # light red  → SPAM
    green_fill = PatternFill("solid", fgColor="D6F5D6")  # light green → NOT SPAM
    red_font   = Font(name="Arial", size=9, color="C0392B", bold=True)
    green_font = Font(name="Arial", size=9, color="1E8449", bold=True)
    base_font  = Font(name="Arial", size=9)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        spam_cell = row[spam_col - 1]
        is_spam   = spam_cell.value
        row_fill  = red_fill if is_spam == "YES" else green_fill

        for cell in row:
            cell.fill = row_fill
            cell.font = base_font

        spam_cell.font      = red_font if is_spam == "YES" else green_font
        spam_cell.alignment = Alignment(horizontal="center")
        row[reason_col - 1].alignment = Alignment(horizontal="center")

    # Column widths
    ws.column_dimensions[get_column_letter(spam_col)].width   = 12
    ws.column_dimensions[get_column_letter(reason_col)].width = 16
    for i, col in enumerate(ws.columns, 1):
        if i not in (spam_col, reason_col):
            ws.column_dimensions[get_column_letter(i)].width = min(
                max(len(str(col[0].value or "")) + 2, 12), 40
            )

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30
    wb.save(output_path)


# =============================================================================
# MAIN
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="CAKE VPBank Social Listening Spam Filter"
    )
    parser.add_argument("--input",  required=True, help="Đường dẫn file Excel đầu vào (.xlsx)")
    parser.add_argument("--output", default=None,  help="Đường dẫn file Excel đầu ra (tuỳ chọn)")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"[ERROR] Không tìm thấy file: {input_path}")
        sys.exit(1)

    output_path = args.output or str(input_path.parent / f"{input_path.stem}_filtered.xlsx")

    print(f"[INFO] Đọc file: {input_path}")
    df = pd.read_excel(input_path)
    print(f"[INFO] Tổng số dòng: {len(df)}")

    # Validate required columns
    missing = [c for c in ("Title", "Content", "Description") if c not in df.columns]
    if missing:
        print(f"[WARNING] Thiếu cột: {missing} — sẽ bỏ qua và dùng chuỗi rỗng")

    df = run_filter(df)

    spam_count    = (df["is_spam"] == "YES").sum()
    nonspam_count = (df["is_spam"] == "NO").sum()

    print(f"[INFO] SPAM=YES : {spam_count}")
    print(f"[INFO] SPAM=NO  : {nonspam_count}")
    print("\n--- Chi tiết phân loại ---")
    print(df.groupby(["is_spam", "spam_reason"]).size().to_string())

    print(f"\n[INFO] Lưu kết quả ra: {output_path}")
    save_colored_excel(df, output_path)
    print("[DONE] Hoàn tất!")


if __name__ == "__main__":
    main()